"""
生成「融合定位组 HF 跟踪」文档。

数据源：
  1. /home/songshu/下载/割草机融合模块版本_20260417_221229.xlsx
     - Sheet1: module_fusion 仓库 commit
     - Sheet2: 其他 repo (slam_workspace / slam_common / plugins) commit
  2. /home/songshu/下载/激光割草机slam重点合入问题跟踪_2026-04-17-22-11-25/
       激光割草机slam重点合入问题跟踪.md

输出：
  /home/songshu/repo/LocalizationTeam/docs/hf-tracking/融合定位组HF跟踪.md

聚合策略：
  - 按 bug 聚合（同一个 bugid 的多次 commit 合成一行）
  - 一个 commit 可能在多个机型有 bugid，每个 (机型, bugid) 拆一行后再按 bugid 聚合
  - 一个 bug 出现在多个机型时，「涉及机型」列出该 bug 出现的机型集合
  - 时间窗口：最近 4 周（2026.3.23 - 2026.4.19）
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/home/songshu/下载/割草机融合模块版本_20260417_221229.xlsx")
MD_PATH = Path(
    "/home/songshu/下载/激光割草机slam重点合入问题跟踪_2026-04-17-22-11-25/"
    "激光割草机slam重点合入问题跟踪.md"
)
OUT_PATH = Path(
    "/home/songshu/repo/LocalizationTeam/docs/hf-tracking/融合定位组HF跟踪.md"
)

MACHINES = ["Butchart", "ButchartPro", "Monet", "Versa"]
ALL_MACHINES_LABEL = "Butchart/ButchartPro/Monet/Versa"

WEEKS = [
    ("2026年第16周（2026.4.13-2026.4.19）", date(2026, 4, 13), date(2026, 4, 19)),
    ("2026年第15周（2026.4.6-2026.4.12）", date(2026, 4, 6), date(2026, 4, 12)),
    ("2026年第14周（2026.3.30-2026.4.5）", date(2026, 3, 30), date(2026, 4, 5)),
    ("2026年第13周（2026.3.23-2026.3.29）", date(2026, 3, 23), date(2026, 3, 29)),
]
WINDOW_START = WEEKS[-1][1]
WINDOW_END = WEEKS[0][2]


@dataclass
class BugRecord:
    """一个逻辑 bug = 一个修复点。

    跨机型时各机型有自己的 bugid（chandao 习惯），都收录到 bug_ids 中。
    """

    key: str  # 聚合键
    bug_ids: dict[str, str] = field(default_factory=dict)  # machine -> bugid
    descriptions: list[str] = field(default_factory=list)
    machines: set[str] = field(default_factory=set)
    commit_dates: list[date] = field(default_factory=list)
    cross_module: str = "-"
    module: str = ""

    def merge(
        self,
        desc: str,
        machine: str,
        bug_id: str,
        commit_date: date,
        cross_module: str = "",
        module: str = "",
    ) -> None:
        if desc and desc not in self.descriptions:
            self.descriptions.append(desc)
        if machine:
            self.machines.add(machine)
            if bug_id and machine not in self.bug_ids:
                self.bug_ids[machine] = bug_id
        if commit_date:
            self.commit_dates.append(commit_date)
        if cross_module and self.cross_module in ("-", ""):
            self.cross_module = cross_module
        if module and not self.module:
            self.module = module

    @property
    def bug_id_label(self) -> str:
        """渲染 bugid 列：单机型直接显示；多机型按 `Butchart:xxx / Monet:yyy` 显示。"""
        if not self.bug_ids:
            return "-"
        if len(self.bug_ids) == 1:
            return next(iter(self.bug_ids.values()))
        ordered = [
            f"{m}: {self.bug_ids[m]}" for m in MACHINES if m in self.bug_ids
        ]
        return "<br>".join(ordered)

    @property
    def first_date(self) -> date:
        return min(self.commit_dates) if self.commit_dates else date.min

    @property
    def last_date(self) -> date:
        return max(self.commit_dates) if self.commit_dates else date.min

    @property
    def machine_label(self) -> str:
        if self.machines == set(MACHINES):
            return ALL_MACHINES_LABEL
        ordered = [m for m in MACHINES if m in self.machines]
        return "/".join(ordered) if ordered else "-"

    @property
    def description(self) -> str:
        return "；".join(self.descriptions)


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    s = s.replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%-d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    parts = s.split("-")
    if len(parts) == 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except ValueError:
            return None
    return None


def normalize_bugid(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s in ("", "-", "—"):
        return ""
    if isinstance(value, float) and value.is_integer():
        s = str(int(value))
    if s.endswith(".0"):
        s = s[:-2]
    return s


def in_window(d: date) -> bool:
    return WINDOW_START <= d <= WINDOW_END


# ---- xlsx 解析 -------------------------------------------------------------


def load_xlsx_records() -> list[dict]:
    """将 xlsx 摊平成行，每行包含一次 commit 在所有机型上的 bugid。

    返回字段：
      - commit_key: 聚合键（同一逻辑 bug 在不同 commit 间合并）
      - date, desc, module
      - per_machine: dict[machine, bugid]
    """
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    records: list[dict] = []

    # Sheet1: module_fusion，列布局：
    #   A=hash, B=comment, C=合入时间, D=作者, E=文档,
    #   F=Butchart, G=Butchart_tag, H=ButchartPro, I=ButchartPro_tag,
    #   J=Monet, K=Monet_tag, L=Versa, M=Versa_tag
    sheet1_machine_cols = {
        "Butchart": 5,
        "ButchartPro": 7,
        "Monet": 9,
        "Versa": 11,
    }
    ws1 = wb["Sheet1"]
    for row in ws1.iter_rows(min_row=3, values_only=True):
        comment = row[1]
        commit_date = parse_date(row[2])
        if not commit_date or not in_window(commit_date):
            continue
        per_machine: dict[str, str] = {}
        for machine, col in sheet1_machine_cols.items():
            bugid = normalize_bugid(row[col])
            if bugid:
                per_machine[machine] = bugid
        if not per_machine:
            continue
        records.append(
            {
                "date": commit_date,
                "desc": (str(comment).strip() if comment else ""),
                "module": "融合(module_fusion)",
                "cross_module": "",
                "per_machine": per_machine,
            }
        )

    # Sheet2: 其他 repo，列布局：
    #   A=hash, B=主要repo, C=comment, D=合入时间, E=作者, F=文档,
    #   G=Butchart, H=ButchartPro, I=Monet, J=Versa
    sheet2_machine_cols = {
        "Butchart": 6,
        "ButchartPro": 7,
        "Monet": 8,
        "Versa": 9,
    }
    ws2 = wb["Sheet2"]
    for row in ws2.iter_rows(min_row=3, values_only=True):
        repo = row[1] or ""
        comment = row[2]
        commit_date = parse_date(row[3])
        if not commit_date or not in_window(commit_date):
            continue
        per_machine = {}
        for machine, col in sheet2_machine_cols.items():
            bugid = normalize_bugid(row[col])
            if bugid:
                per_machine[machine] = bugid
        if not per_machine:
            continue
        records.append(
            {
                "date": commit_date,
                "desc": (str(comment).strip() if comment else ""),
                "module": f"其他({repo})" if repo else "其他",
                "cross_module": "",
                "per_machine": per_machine,
            }
        )

    return records


# ---- md 解析 ---------------------------------------------------------------


MD_ROW_RE = re.compile(r"^\|(.+)\|$")


def split_md_row(line: str) -> list[str] | None:
    m = MD_ROW_RE.match(line.strip())
    if not m:
        return None
    parts = [c.strip() for c in m.group(1).split("|")]
    return parts


def load_md_records() -> list[dict]:
    """解析重点合入问题跟踪 md 表格。"""
    records: list[dict] = []
    text = MD_PATH.read_text(encoding="utf-8")
    in_table = False
    headers: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line.startswith("|"):
            in_table = False
            continue
        cells = split_md_row(line)
        if cells is None:
            continue
        if all(set(c) <= set("-: ") for c in cells):
            in_table = True
            continue
        if not in_table:
            headers = cells
            continue

        def col(name: str) -> str:
            try:
                idx = headers.index(name)
            except ValueError:
                return ""
            return cells[idx] if idx < len(cells) else ""

        # 去掉删除线 ~~ ... ~~ 的整行（已废弃记录）
        joined = "".join(cells)
        if joined.count("~~") >= 2 and joined.startswith("~~"):
            continue

        bugid = normalize_bugid(col("问题ID"))
        if not bugid:
            continue
        commit_date = parse_date(col("合入时间"))
        if not commit_date or not in_window(commit_date):
            continue
        machines_raw = col("涉及机型")
        machines = parse_machines_field(machines_raw)
        desc = col("问题描述")
        cross = col("是否影响其他模块") or "-"
        module = col("模块") or ""
        # md 表里一个 bugid 对应多个机型，所有机型共用一个 bugid
        per_machine = {m: bugid for m in machines} if machines else {}
        records.append(
            {
                "date": commit_date,
                "desc": desc,
                "module": module,
                "cross_module": cross,
                "per_machine": per_machine,
                "primary_bug_id": bugid,
            }
        )
    return records


def parse_machines_field(raw: str) -> list[str]:
    if not raw:
        return []
    s = raw.replace(" ", "").replace("、", "/").replace(",", "/")
    if s == "全部":
        return list(MACHINES)
    out = []
    for token in s.split("/"):
        for m in MACHINES:
            if token.lower() == m.lower() and m not in out:
                out.append(m)
                break
    return out


# ---- 聚合与渲染 -----------------------------------------------------------


def make_key(record: dict) -> str:
    """聚合键：

    优先按 bugid 集合（任意一个 bugid 命中已存在 record 即合并）。
    简化：用 record 中第一个机型的 bugid 作为主键；后续在 aggregate 中用
    bugid 索引合并。
    """
    pm = record.get("per_machine", {})
    if not pm:
        return f"desc::{record['desc'].strip().lower()}::{record['date']}"
    # 用所有 bugid 排序拼成 key
    return "|".join(sorted(pm.values()))


def aggregate(records: list[dict]) -> list[BugRecord]:
    """按 bugid 集合聚合。

    任意一个 bugid 在已有 BugRecord 中出现，就并入；否则新建。
    """
    bugs: list[BugRecord] = []
    bugid_index: dict[str, BugRecord] = {}

    for r in records:
        pm = r.get("per_machine", {})
        target: BugRecord | None = None
        for bid in pm.values():
            if bid in bugid_index:
                target = bugid_index[bid]
                break
        if target is None:
            key = make_key(r)
            target = BugRecord(key=key)
            bugs.append(target)
        for machine, bid in pm.items():
            target.merge(
                desc=r["desc"],
                machine=machine,
                bug_id=bid,
                commit_date=r["date"],
                cross_module=r["cross_module"],
                module=r["module"],
            )
            bugid_index[bid] = target
    return bugs


def assign_to_weeks(bugs: list[BugRecord]) -> dict[str, list[BugRecord]]:
    """按「最早 commit 时间」归到所在周。"""
    grouped: dict[str, list[BugRecord]] = defaultdict(list)
    for bug in bugs:
        d = bug.first_date
        for label, start, end in WEEKS:
            if start <= d <= end:
                grouped[label].append(bug)
                break
    for label in grouped:
        grouped[label].sort(key=lambda b: (b.first_date, b.key))
    return grouped


def render(grouped: dict[str, list[BugRecord]]) -> str:
    lines = []
    lines.append("# 融合定位组 HF 跟踪")
    lines.append("")
    lines.append(
        f"> 数据源：`module_fusion` 全量 commit + 激光 SLAM 重点合入跟踪表  "
    )
    lines.append(
        f"> 时间窗口：{WINDOW_START.isoformat()} ~ {WINDOW_END.isoformat()}（最近 4 周）  "
    )
    lines.append(
        f"> 生成方式：`scripts/build_hf_tracking.py`，按 bug 聚合，"
        f"同一 bug 的多次 commit 取首次合入时间"
    )
    lines.append("")

    for label, _start, _end in WEEKS:
        lines.append(f"## {label}")
        lines.append("")
        bugs = grouped.get(label, [])
        if not bugs:
            lines.append("_本周无合入记录_")
            lines.append("")
            continue
        lines.append(
            "| 合入 dev 时间 | bug id | 功能 / 优化问题 | 涉及机型 | 是否涉及其他模块 |"
        )
        lines.append("| --- | --- | --- | --- | --- |")
        for bug in bugs:
            date_str = bug.first_date.isoformat()
            if bug.last_date != bug.first_date:
                date_str = f"{bug.first_date.isoformat()} ~ {bug.last_date.isoformat()}"
            desc = bug.description.replace("|", "\\|").replace("\n", " ")
            lines.append(
                f"| {date_str} | {bug.bug_id_label} | {desc} | "
                f"{bug.machine_label} | {bug.cross_module} |"
            )
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    xlsx_records = load_xlsx_records()
    md_records = load_md_records()
    all_records = xlsx_records + md_records
    bugs = aggregate(all_records)
    grouped = assign_to_weeks(bugs)
    md = render(grouped)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(md, encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    total = sum(len(v) for v in grouped.values())
    print(f"Total bugs in 4 weeks: {total}")
    for label, _s, _e in WEEKS:
        print(f"  {label}: {len(grouped.get(label, []))}")


if __name__ == "__main__":
    main()
